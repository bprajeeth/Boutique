from django.shortcuts import render
from django.http import HttpResponse
from django.shortcuts import render
from django.http import JsonResponse
from client.models import *
from openpyxl import Workbook
import json
# from win32com import client as x
import os
import pywhatkit
# Create your 
# views here.
# import  jpype     
# import  asposecells
# from asposecells.api import Workbook

def home_view(request):
    return render(request, "Home.html")


def form_view(request):
    if request.method == "POST":
        print(request.POST)
        # VALUE INSIDE BRACKETS ARE ID FROM HTML TAGS
        username = request.POST["name"]
        mob = request.POST["mobile"]
        dress = request.POST["dress"]

        shoulder = request.POST["shoulder"]
        bust1 = request.POST["bust1"]
        bust2 = request.POST["bust2"]
        bust3 = request.POST["bust3"]
        blouse = request.POST["blouse"]
        arm1 = request.POST["arm1"]
        arm2 = request.POST["arm2"]
        arm3 = request.POST["arm3"]
        wst = request.POST["waist"]
        saree1 = request.POST["saree1"]
        saree2 = request.POST["saree2"]
        knee = request.POST["knee"]
        thav = request.POST["thavani"]
        pyjama1 = request.POST["pyjama1"]
        pyjama2 = request.POST["pyjama2"]
        pyjama3 = request.POST["pyjama3"]
        pyjama4 = request.POST["pyjama4"]

        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1).value = "name"
        ws.cell(row=1, column=2).value = username
        ws.cell(row=2, column=1).value = "mobile"
        ws.cell(row=2, column=2).value = mob
        ws.cell(row=3, column=1).value = "dress_type"
        ws.cell(row=3, column=2).value = dress
        ws.cell(row=4, column=2).value = ""
        ws.cell(row=5, column=2).value = "MEASUREMENTS"
        ws.cell(row=6, column=2).value = ""
        ws.cell(row=7, column=1).value = "shoulder_to_shoulder"
        ws.cell(row=7, column=2).value = shoulder
        ws.cell(row=8, column=1).value = "above_bust"
        ws.cell(row=8, column=2).value = bust1
        ws.cell(row=9, column=1).value = "top_bust"
        ws.cell(row=9, column=2).value = bust2
        ws.cell(row=10, column=1).value = "below_bust"
        ws.cell(row=10, column=2).value = bust3
        ws.cell(row=11, column=1).value = "blouse_length"
        ws.cell(row=11, column=2).value = blouse
        ws.cell(row=12, column=1).value = "arm_hole"
        ws.cell(row=12, column=2).value = arm1
        ws.cell(row=13, column=1).value = "arm_width"
        ws.cell(row=13, column=2).value = arm2
        ws.cell(row=14, column=1).value = "arm_length"
        ws.cell(row=14, column=2).value = arm3
        ws.cell(row=15, column=1).value = "waist"
        ws.cell(row=15, column=2).value = wst
        ws.cell(row=16, column=1).value = "saree_length"
        ws.cell(row=17, column=2).value = saree1
        ws.cell(row=18, column=1).value = "saree_waist"
        ws.cell(row=18, column=2).value = saree2
        ws.cell(row=19, column=1).value = "knee_to_knee"
        ws.cell(row=19, column=2).value = knee
        ws.cell(row=20, column=1).value = "thavani"
        ws.cell(row=20, column=2).value = thav
        ws.cell(row=21, column=1).value = "pyjama_waist"
        ws.cell(row=21, column=2).value = pyjama1
        ws.cell(row=22, column=1).value = "pyjama_hip"
        ws.cell(row=22, column=2).value = pyjama2
        ws.cell(row=23, column=1).value = "pyjama_length"
        ws.cell(row=23, column=2).value = pyjama3
        ws.cell(row=24, column=1).value = "pyjama_ankle"
        ws.cell(row=24, column=2).value = pyjama4
        wb.save("check.xlsx")
        

        user = client(
            name=username,
            mobile=mob,
            dress_type=dress,
            shoulder_to_shoulder=shoulder,
            above_bust=bust1,
            top_bust=bust2,
            below_bust=bust3,
            blouse_length=blouse,
            arm_hole=arm1,
            arm_width=arm2,
            arm_length=arm3,
            waist=wst,
            saree_length=saree1,
            saree_waist=saree2,
            knee_to_knee=knee,
            thavani=thav,
            pyjama_waist=pyjama1,
            pyjama_hip=pyjama2,
            pyjama_length=pyjama3,
            pyjama_ankle=pyjama4,
        )
        # user = client(shoulder_to_shoulder=shoulder, above_bust=bust1, top_bust=bust2 ,below_bust=bust3 , blouse_length=blouse , arm_hole=arm1, arm_width=arm2,  arm_length=arm3, waist=wst , saree_length=saree1, saree_waist=saree2, knee_to_knee=knee , thavani=thav, pyjama_waist=pyjama1, pyjama_hip=pyjama2, pyjama_length=pyjama3, pyjama_ankle=pyjama4)
        user.save()
     
        # jpype.startJVM() 
        # workbook = Workbook("check.xlsx")
        # workbook.Save("Output.png")
        # jpype.shutdownJVM()
        pywhatkit.sendwhatmsg_instantly("+91"+mob,username+" -Form-Sent")
        # os.remove("check.xlsx")
        # # Read Excel File
        # sheets = excel.Workbooks.Open("check.xlsx")
        # work_sheets = sheets.Worksheets[0]
        # # Convert into PDF File
        # work_sheets.ExportAsFixedFormat(0, "checkpdf.pdf")

        # workbook = Workbook("check.xlsx")
        # Convert Excel to PDF
        #workbook.save("xlsx-to-pdf.pdf", SaveFormat.PDF)
                
        # excel = x.Dispatch("Excel.Application")
        
        # # Read Excel File
        # sheets = excel.Workbooks.Open('check.xlsx')
        # work_sheets = sheets.Worksheets[0]
        
        # # Convert into PDF File
        # work_sheets.ActiveSheet.ExportAsFixedFormat(0, 'check.pdf')
        # sheets.Close()        
        
        return HttpResponse("succesfull")
    else:
        return render(request, "form.html")


def costumes_view(request):
    pattern=pattern_display.objects.all()
    return render(request, "Patterns.html", {'costume':pattern} )
    # return HttpResponse("Patterns.html")


def ornaments_view(request):
    ornaments=ornaments_display.objects.all()
    return render(request, "Jewellery.html", {"ornament":ornaments})


def accessories_view(request):
    accessories=accessories_display.objects.all()
    return render(request, "Accessories.html",{"accessories":accessories})

def get_jewellery(request):
    posts=list(ornaments_display.objects.values())
    print("posts = ",posts)
    return JsonResponse({'data':posts},safe=False)



